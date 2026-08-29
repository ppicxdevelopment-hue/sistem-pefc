import streamlit as st
import pandas as pd
import math
import os
import io
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# ==========================================
# 0. SISTEM KEAMANAN PIN PABRIK
# ==========================================
# Silakan ganti angka ini dengan PIN rahasia untuk tim Anda
PIN_RAHASIA = "2026" 

if "akses_diberikan" not in st.session_state:
    st.session_state["akses_diberikan"] = False

if not st.session_state["akses_diberikan"]:
    st.title("🔒 Area Terbatas Pabrik")
    st.info("Silakan masukkan PIN akses untuk menggunakan sistem Request Kode PEFC.")
    
    pin_input = st.text_input("Masukkan PIN:", type="password")
    
    if st.button("Buka Kunci Sistem", type="primary"):
        if pin_input == PIN_RAHASIA:
            st.session_state["akses_diberikan"] = True
            st.rerun()
        else:
            st.error("❌ PIN Salah! Akses ditolak.")
            
    # Kode st.stop() ini sangat penting agar kode di bawahnya tidak dijalankan
    # sampai PIN yang dimasukkan benar.
    st.stop() 

# ==========================================
# 1. KONFIGURASI KONEKSI GOOGLE SHEETS
# ==========================================
st.set_page_config(page_title="Generator Kode PEFC", layout="wide")
st.title("🏭 Sistem Request Kode PEFC (Online Google Sheets)")


import streamlit as st
import pandas as pd
import math
import os
import io
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# ==========================================
# 1. KONFIGURASI KONEKSI GOOGLE SHEETS
# ==========================================
st.set_page_config(page_title="Generator Kode PEFC", layout="wide")
st.title("🏭 Sistem Request Kode PEFC (Online Google Sheets)")

CREDENTIALS_PATH = r"D:\SERLYA\PROYEK\SISTEM PEFC\credentials.json"

SHEET_STOCK = 'Master_stock PEFC'
SHEET_KOMPOSISI = 'Master_komposisi PEFC'
SHEET_LOG = 'log_transaksi'
SHEET_DETAIL = 'log_detail_bahan'

def get_gsheet_client():
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
        "https://www.googleapis.com/auth/drive"
    ]
    try:
        if "GOOGLE_CREDENTIALS" in st.secrets:
            creds_dict = json.loads(st.secrets["GOOGLE_CREDENTIALS"])
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        else:
            creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_PATH, scope)
        client = gspread.authorize(creds)
        return client
    except Exception as e:
        st.error(f"❌ Gagal koneksi ke Google API: {e}")
        return None

@st.cache_data(ttl=30)
def load_data_gspread():
    client = get_gsheet_client()
    if not client: return pd.DataFrame(), pd.DataFrame()
    try:
        sheet_s = client.open(SHEET_STOCK).sheet1
        df_stok = pd.DataFrame(sheet_s.get_all_records())
        if 'Tgl Kedatangan' in df_stok.columns:
            df_stok['Tgl Kedatangan'] = pd.to_datetime(df_stok['Tgl Kedatangan'], errors='coerce')
        sheet_k = client.open(SHEET_KOMPOSISI).sheet1
        df_komposisi = pd.DataFrame(sheet_k.get_all_records())
        return df_stok, df_komposisi
    except Exception as e:
        st.error(f"❌ Gagal membaca Google Sheets: {e}")
        return pd.DataFrame(), pd.DataFrame()

df_stok_master, df_komposisi = load_data_gspread()

def save_stock_gspread(df_stok):
    client = get_gsheet_client()
    if not client: return
    sheet_s = client.open(SHEET_STOCK).sheet1
    df_upload = df_stok.copy()
    if 'Tgl Kedatangan' in df_upload.columns:
        df_upload['Tgl Kedatangan'] = pd.to_datetime(df_upload['Tgl Kedatangan'], errors='coerce').dt.strftime('%Y-%m-%d')
    sheet_s.clear()
    sheet_s.update([df_upload.columns.values.tolist()] + df_upload.fillna("").values.tolist())

def append_log_gspread_strict(sheet_name, data_dict_list):
    client = get_gsheet_client()
    if not client: return
    sheet = client.open(sheet_name).sheet1
    for row_data in data_dict_list:
        if sheet_name == SHEET_LOG:
            row_values = [
                str(row_data.get("No_PRO", "")), str(row_data.get("Tanggal", "")),
                str(row_data.get("Requester", "")), str(row_data.get("Customer", "")),
                str(row_data.get("Kode_PEFC", "")), float(row_data.get("Tonase", 0.0))
            ]
        elif sheet_name == SHEET_DETAIL:
            row_values = [
                str(row_data.get("Tipe_Pulp", "")), str(row_data.get("Batch", "")),
                str(row_data.get("Desc", "")), str(row_data.get("Kode_Bahan", "")),
                float(row_data.get("Kg_Terpakai", 0.0)), int(row_data.get("Bale_Terpakai", 0)),
                str(row_data.get("No_PRO", "")), str(row_data.get("Customer", "")),
                str(row_data.get("Kode_PEFC", ""))
            ]
        else:
            row_values = list(row_data.values())
        sheet.append_row(row_values)

@st.cache_data(ttl=30)
def get_all_log_gspread(sheet_name):
    client = get_gsheet_client()
    if not client: return pd.DataFrame()
    try:
        sheet = client.open(sheet_name).sheet1
        data = sheet.get_all_records()
        return pd.DataFrame(data)
    except:
        return pd.DataFrame()

# ==========================================
# 2. FUNGSI KALKULASI & ALOKASI
# ==========================================
def parse_percentage(val):
    if pd.isna(val): return 0.0
    if isinstance(val, str):
        val = val.replace('%', '').strip()
        try:
            val_float = float(val)
            return val_float / 100.0 if val_float > 1 else val_float
        except:
            return 0.0
    return float(val)

def get_rasio_komposisi(grade_input, claim_input):
    mapping_grade = {"SANITARY NAPKIN": "SN", "TOWEL": "TW", "FACIAL": "FACIAL", "MG": "MG", "NAPKIN": "N", "TOILET": "TOILET"}
    g_in = str(grade_input).strip().upper()
    grade_excel = mapping_grade.get(g_in, g_in)
    c_in = str(claim_input).strip().upper()
    df_temp = df_komposisi.copy()
    df_temp.columns = df_temp.columns.astype(str).str.strip().str.upper()
    df_temp['GRADE_CLEAN'] = df_temp['GRADE'].astype(str).str.strip().str.upper()
    df_temp['CLAIM_CLEAN'] = df_temp['CLAIM PEFC'].astype(str).str.strip().str.upper()
    komp_match = df_temp[(df_temp['GRADE_CLEAN'] == grade_excel) & (df_temp['CLAIM_CLEAN'] == c_in)]
    if not komp_match.empty:
        nbkp_ctrl = parse_percentage(komp_match['NBKP_CONTROLLED'].values[0]) if 'NBKP_CONTROLLED' in df_temp.columns else 0.0
        nbkp_pefc = parse_percentage(komp_match['NBKP_PEFC'].values[0]) if 'NBKP_PEFC' in df_temp.columns else 0.0
        ebkp_pefc = parse_percentage(komp_match['EBKP_PEFC'].values[0]) if 'EBKP_PEFC' in df_temp.columns else 0.0
        return nbkp_ctrl, nbkp_pefc, ebkp_pefc
    return 0.0, 0.30, 0.70

def get_next_sequence(tahun_bulan):
    df_log = get_all_log_gspread(SHEET_LOG)
    if df_log.empty or 'Tanggal' not in df_log.columns: return "001"
    df_log['YYMM'] = pd.to_datetime(df_log['Tanggal'], errors='coerce').dt.strftime('%y%m')
    return str(len(df_log[df_log['YYMM'] == tahun_bulan]) + 1).zfill(3)

def alokasi_fifo(kebutuhan_kg, df_stok, tipe_material, is_eudr, filter_claim_bb=None):
    if kebutuhan_kg <= 0: return [], 0.0
    stok_valid = df_stok[df_stok['Material type'] == tipe_material].copy()
    if is_eudr == 'Ya':
        stok_valid = stok_valid[stok_valid['CLAIM EUDR'].astype(str).str.strip().str.upper() == 'EUDR']
    if filter_claim_bb:
        stok_valid = stok_valid[stok_valid['Claim BB'].astype(str).str.contains(filter_claim_bb, case=False, na=False)]
    stok_valid = stok_valid.sort_values(by='Tgl Kedatangan', ascending=True)
    alokasi_batch = []
    sisa_kebutuhan = kebutuhan_kg
    for idx, row in stok_valid.iterrows():
        if sisa_kebutuhan <= 0: break
        qty_val = float(row['Quantity (kg)']) if pd.notna(row['Quantity (kg)']) else 0.0
        if qty_val <= 0: continue
        ambil_kg = min(sisa_kebutuhan, qty_val)
        sisa_kebutuhan -= ambil_kg
        konversi_bale = float(row['Conv (kg/bale)']) if pd.notna(row['Conv (kg/bale)']) else 0.0
        bale_diambil = math.ceil(ambil_kg / konversi_bale) if konversi_bale > 0 else 0
        alokasi_batch.append({
            'Tipe_Pulp': tipe_material, 'Batch': row['Batch'], 'Desc': row['Batch Description'],
            'Kode_Bahan': row['FSC CODE'], 'Kg_Terpakai': ambil_kg, 'Bale_Terpakai': bale_diambil
        })
    return alokasi_batch, sisa_kebutuhan

def alokasi_fifo_constrained(kebutuhan_kg, df_stok, tipe_material, allowed_codes):
    stok_valid = df_stok[(df_stok['Material type'] == tipe_material) & (df_stok['FSC CODE'].isin(allowed_codes))].copy()
    stok_valid = stok_valid.sort_values(by='Tgl Kedatangan', ascending=True)
    alokasi_batch = []
    sisa_kebutuhan = kebutuhan_kg
    for idx, row in stok_valid.iterrows():
        if sisa_kebutuhan <= 0: break
        qty_val = float(row['Quantity (kg)']) if pd.notna(row['Quantity (kg)']) else 0.0
        if qty_val <= 0: continue
        ambil_kg = min(sisa_kebutuhan, qty_val)
        sisa_kebutuhan -= ambil_kg
        konversi_bale = float(row['Conv (kg/bale)']) if pd.notna(row['Conv (kg/bale)']) else 0.0
        bale_diambil = math.ceil(ambil_kg / konversi_bale) if konversi_bale > 0 else 0
        alokasi_batch.append({
            'Tipe_Pulp': tipe_material, 'Batch': row['Batch'], 'Desc': row['Batch Description'],
            'Kode_Bahan': row['FSC CODE'], 'Kg_Terpakai': ambil_kg, 'Bale_Terpakai': bale_diambil
        })
    return alokasi_batch, sisa_kebutuhan

# ==========================================
# 3. ANTARMUKA TAB LENGKAP (6 TAB)
# ==========================================
tab_request, tab_report, tab_revisi, tab_stok_masuk, tab_dashboard, tab_admin = st.tabs([
    "📝 Request PRO", "🔍 Report & Export", "✏️ Revisi PRO", "📦 Terima Stok", "📊 Dashboard Stok", "⚙️ Admin & Reset"
])

# ----------------- TAB 1: REQUEST -----------------
with tab_request:
    with st.sidebar:
        st.header("📋 Form Input PRO")
        with st.form("form_request_pro"):
            no_pro = st.text_input("No PRO (Order Produksi)", placeholder="Contoh: PRO-8821")
            tgl_pro = st.date_input("Tgl PRO")
            requester = st.text_input("Nama Requester")
            customer = st.text_input("Customer")
            nama_mesin = st.selectbox("Nama Mesin", [f"PM{i}" for i in range(11, 26)])
            grade = st.selectbox("Grade", ["Facial", "Sanitary Napkin", "MG", "Napkin", "Toilet", "Towel"])
            claim_pefc = st.selectbox("Claim PEFC", ["PEFC CERTIFIED", "70% PEFC CERTIFIED"])
            req_tonase = st.number_input("Request Tonase (Kg)", min_value=1.0, value=1000.0, step=100.0)
            claim_eudr = st.radio("Claim EUDR?", ["Tidak", "Ya"])
            btn_generate = st.form_submit_button("🚀 Generate Kode PEFC", type="primary")

    if btn_generate:
        if no_pro == "":
            st.warning("⚠️ Harap isi 'No PRO' terlebih dahulu.")
        else:
            with st.spinner("⏳ Menghubungkan ke Google Sheets & memproses alokasi..."):
                df_log_cek = get_all_log_gspread(SHEET_LOG)
                is_duplicate = False
                if not df_log_cek.empty and 'No_PRO' in df_log_cek.columns:
                    if str(no_pro) in df_log_cek['No_PRO'].astype(str).values: is_duplicate = True
                
                if is_duplicate:
                    st.error(f"❌ Gagal! No PRO '{no_pro}' sudah pernah digenerate.")
                elif not df_stok_master.empty:
                    rasio_nbkp_ctrl, rasio_nbkp_pefc, rasio_ebkp_pefc = get_rasio_komposisi(grade, claim_pefc)
                    total_kebutuhan_pulp = req_tonase / 0.93
                    
                    hasil_nbkp_ctrl, sisa_nbkp_ctrl = alokasi_fifo(total_kebutuhan_pulp * rasio_nbkp_ctrl, df_stok_master, "NBKP", claim_eudr, "Controlled")
                    hasil_nbkp_pefc, sisa_nbkp_pefc = alokasi_fifo(total_kebutuhan_pulp * rasio_nbkp_pefc, df_stok_master, "NBKP", claim_eudr, "PEFC")
                    hasil_ebkp_pefc, sisa_ebkp_pefc = alokasi_fifo(total_kebutuhan_pulp * rasio_ebkp_pefc, df_stok_master, "EBKP", claim_eudr, "PEFC")
                    
                    if sisa_nbkp_ctrl > 0.1 or sisa_nbkp_pefc > 0.1 or sisa_ebkp_pefc > 0.1:
                        st.error(f"⚠️ Stok Google Sheets tidak mencukupi!\nKurang NBKP (Controlled): {sisa_nbkp_ctrl:.2f} Kg\nKurang NBKP (PEFC): {sisa_nbkp_pefc:.2f} Kg\nKurang EBKP (PEFC): {sisa_ebkp_pefc:.2f} Kg")
                    else:
                        semua_alokasi = hasil_nbkp_ctrl + hasil_nbkp_pefc + hasil_ebkp_pefc
                        kode_bahan_terpakai = list(set([item['Kode_Bahan'] for item in semua_alokasi if pd.notna(item['Kode_Bahan'])]))
                        
                        if len(kode_bahan_terpakai) > 4:
                            st.error("⚠️ Kombinasi bahan baku melebihi 4 jenis batch/kode!")
                        else:
                            kode_bahan_terpakai.sort()
                            while len(kode_bahan_terpakai) < 4: kode_bahan_terpakai.append("00R")
                            nomor_urut = get_next_sequence(tgl_pro.strftime('%y%m'))
                            kode_pefc_final = f"{kode_bahan_terpakai[0]}-{kode_bahan_terpakai[1]}-{kode_bahan_terpakai[2]}-{kode_bahan_terpakai[3]}-{nomor_urut}{tgl_pro.strftime('%y')}"
                            
                            st.success(f"**KODE PEFC GENERATED (ONLINE):** {kode_pefc_final}")
                            try:
                                for item in semua_alokasi:
                                    batch_mask = df_stok_master['Batch'].astype(str) == str(item['Batch'])
                                    current_q = float(df_stok_master.loc[batch_mask, 'Quantity (kg)'].values[0])
                                    current_b = float(df_stok_master.loc[batch_mask, 'Qty BAL'].values[0])
                                    df_stok_master.loc[batch_mask, 'Quantity (kg)'] = max(0.0, current_q - item['Kg_Terpakai'])
                                    df_stok_master.loc[batch_mask, 'Qty BAL'] = max(0, current_b - item['Bale_Terpakai'])
                                save_stock_gspread(df_stok_master)
                                
                                log_row = [{
                                    "No_PRO": str(no_pro), "Tanggal": tgl_pro.strftime('%Y-%m-%d'),
                                    "Requester": str(requester), "Customer": str(customer),
                                    "Kode_PEFC": str(kode_pefc_final), "Tonase": float(req_tonase)
                                }]
                                append_log_gspread_strict(SHEET_LOG, log_row)
                                
                                detail_rows = []
                                for item in semua_alokasi:
                                    item_copy = item.copy()
                                    item_copy['No_PRO'] = str(no_pro)
                                    item_copy['Customer'] = str(customer)
                                    item_copy['Kode_PEFC'] = str(kode_pefc_final)
                                    detail_rows.append(item_copy)
                                append_log_gspread_strict(SHEET_DETAIL, detail_rows)
                                st.info(f"✨ Mutasi berhasil disimpan secara online ke Google Sheets!")
                            except Exception as e:
                                st.error(f"❌ GAGAL menyimpan ke Google Sheets: {e}")

# ----------------- TAB 2: REPORT & EXPORT -----------------
with tab_report:
    st.header("🔍 Cek Material & Export Laporan Online")
    keyword = st.text_input("Masukkan No PRO / Kode PEFC:", placeholder="Contoh: PRO-8821...")
    if keyword:
        with st.spinner("Mencari data..."):
            df_hist = get_all_log_gspread(SHEET_DETAIL)
            if not df_hist.empty:
                hasil_cari = df_hist[df_hist['No_PRO'].astype(str).str.contains(keyword, case=False, na=False) | df_hist['Kode_PEFC'].astype(str).str.contains(keyword, case=False, na=False)]
                if not hasil_cari.empty:
                    st.success(f"Ditemukan! **No PRO:** {hasil_cari['No_PRO'].iloc[0]} | **Kode PEFC:** {hasil_cari['Kode_PEFC'].iloc[0]}")
                    cols_to_show = ['Tipe_Pulp', 'Kode_Bahan', 'Batch', 'Desc', 'Kg_Terpakai', 'Bale_Terpakai']
                    if 'Customer' in hasil_cari.columns: cols_to_show.insert(0, 'Customer')
                    st.dataframe(hasil_cari[cols_to_show])
                else:
                    st.warning("Data tidak ditemukan.")
                
    st.markdown("---")
    st.subheader("📥 Download Semua Riwayat ke Excel")
    if st.button("📊 Tarik & Download Laporan (.xlsx)"):
        with st.spinner("Mengunduh laporan dari Google Sheets..."):
            df_log_utama = get_all_log_gspread(SHEET_LOG)
            df_log_detail = get_all_log_gspread(SHEET_DETAIL)
            output_report = io.BytesIO()
            with pd.ExcelWriter(output_report, engine='openpyxl') as writer:
                if not df_log_utama.empty: df_log_utama.to_excel(writer, index=False, sheet_name='Summary_PRO')
                if not df_log_detail.empty: df_log_detail.to_excel(writer, index=False, sheet_name='Detail_Bahan_Baku')
            output_report.seek(0)
            st.download_button("💾 Simpan File Excel", data=output_report, file_name=f"Report_Online_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ----------------- TAB 3: REVISI PRO -----------------
with tab_revisi:
    st.header("✏️ Revisi & Penyesuaian Transaksi")
    jenis_revisi = st.selectbox("Pilih Jenis Revisi:", [
        "1. Tambah Kuantiti PRO (Kunci Batch/Kode)",
        "2. Batalkan / Hapus Transaksi PRO", 
        "3. Update Hasil Aktual (Smart Penyesuaian +/- Stok)"
    ])
    rev_no_pro = st.text_input("Masukkan No PRO yang ingin direvisi:", key="input_revisi_pro_clean")
    
    df_log_all = get_all_log_gspread(SHEET_LOG)
    total_req_saat_ini = 0.0
    if rev_no_pro != "" and not df_log_all.empty and 'No_PRO' in df_log_all.columns:
        match_pro = df_log_all[df_log_all['No_PRO'].astype(str) == str(rev_no_pro)]
        if not match_pro.empty:
            total_req_saat_ini = pd.to_numeric(match_pro['Tonase'], errors='coerce').sum()
            st.info(f"📌 **Total Request saat ini untuk {rev_no_pro}:** **{total_req_saat_ini:,.2f} Kg**.")

    if rev_no_pro != "":
        df_detail_all = get_all_log_gspread(SHEET_DETAIL)
        
        if jenis_revisi.startswith("1."):
            tambah_tonase = st.number_input("Tambahan Request Tonase (Kg)", min_value=1.0, value=100.0, step=50.0)
            if st.button("➕ Tambah Kuantiti", type="primary"):
                with st.spinner("⏳ Memproses penambahan kuantiti & update stok..."):
                    hist_pro = df_detail_all[df_detail_all['No_PRO'].astype(str) == str(rev_no_pro)]
                    if hist_pro.empty:
                        st.warning("No PRO tidak ditemukan di detail.")
                    else:
                        allowed_nbkp = hist_pro[hist_pro['Tipe_Pulp'] == 'NBKP']['Kode_Bahan'].unique().tolist()
                        allowed_ebkp = hist_pro[hist_pro['Tipe_Pulp'] == 'EBKP']['Kode_Bahan'].unique().tolist()
                        tot_nbkp = pd.to_numeric(hist_pro[hist_pro['Tipe_Pulp'] == 'NBKP']['Kg_Terpakai'], errors='coerce').sum()
                        tot_ebkp = pd.to_numeric(hist_pro[hist_pro['Tipe_Pulp'] == 'EBKP']['Kg_Terpakai'], errors='coerce').sum()
                        tot_pulp = tot_nbkp + tot_ebkp
                        rasio_nbkp = tot_nbkp / tot_pulp if tot_pulp > 0 else 0
                        rasio_ebkp = tot_ebkp / tot_pulp if tot_pulp > 0 else 0
                        
                        tambahan_pulp = tambah_tonase / 0.93
                        hasil_nbkp, sisa_nbkp = alokasi_fifo_constrained(tambahan_pulp * rasio_nbkp, df_stok_master, "NBKP", allowed_nbkp)
                        hasil_ebkp, sisa_ebkp = alokasi_fifo_constrained(tambahan_pulp * rasio_ebkp, df_stok_master, "EBKP", allowed_ebkp)
                        
                        if sisa_nbkp > 0.1 or sisa_ebkp > 0.1:
                            st.error("❌ Stok Google Sheets dengan kode PEFC yang sama sudah habis!")
                        else:
                            semua_alokasi = hasil_nbkp + hasil_ebkp
                            for item in semua_alokasi:
                                batch_mask = df_stok_master['Batch'].astype(str) == str(item['Batch'])
                                df_stok_master.loc[batch_mask, 'Quantity (kg)'] = max(0.0, float(df_stok_master.loc[batch_mask, 'Quantity (kg)'].values[0]) - item['Kg_Terpakai'])
                                df_stok_master.loc[batch_mask, 'Qty BAL'] = max(0, int(df_stok_master.loc[batch_mask, 'Qty BAL'].values[0]) - item['Bale_Terpakai'])
                            save_stock_gspread(df_stok_master)
                            
                            kode_pefc_lama = hist_pro['Kode_PEFC'].iloc[0]
                            customer_lama = hist_pro['Customer'].iloc[0] if 'Customer' in hist_pro.columns else "Unknown"
                            
                            detail_rows = []
                            for item in semua_alokasi:
                                item_copy = item.copy()
                                item_copy['No_PRO'] = str(rev_no_pro)
                                item_copy['Customer'] = str(customer_lama)
                                item_copy['Kode_PEFC'] = str(kode_pefc_lama)
                                detail_rows.append(item_copy)
                            append_log_gspread_strict(SHEET_DETAIL, detail_rows)
                            
                            log_row = [{
                                "No_PRO": str(rev_no_pro), "Tanggal": datetime.now().strftime('%Y-%m-%d'),
                                "Requester": "System (Revisi)", "Customer": str(customer_lama),
                                "Kode_PEFC": str(kode_pefc_lama), "Tonase": float(tambah_tonase)
                            }]
                            append_log_gspread_strict(SHEET_LOG, log_row)
                            st.success(f"✅ Penambahan {tambah_tonase} Kg berhasil dipotong dari batch yang sama.")

        elif jenis_revisi.startswith("2."):
            if st.button("🗑️ Batalkan PRO & Kembalikan Stok", type="primary"):
                with st.spinner("⏳ Membatalkan transaksi & mencatat riwayat minus..."):
                    hist_pro = df_detail_all[df_detail_all['No_PRO'].astype(str) == str(rev_no_pro)]
                    if hist_pro.empty:
                        st.warning("No PRO tidak ditemukan.")
                    elif total_req_saat_ini <= 0:
                        st.warning("⚠️ PRO ini sudah pernah dibatalkan (Total Request saat ini 0).")
                    else:
                        net_detail = hist_pro.groupby(['Tipe_Pulp', 'Batch', 'Desc', 'Kode_Bahan'])[['Kg_Terpakai', 'Bale_Terpakai']].sum().reset_index()
                        detail_reversals = []
                        for idx, row in net_detail.iterrows():
                            if row['Kg_Terpakai'] > 0:
                                batch_mask = df_stok_master['Batch'].astype(str) == str(row['Batch'])
                                if not df_stok_master[batch_mask].empty:
                                    df_stok_master.loc[batch_mask, 'Quantity (kg)'] = float(df_stok_master.loc[batch_mask, 'Quantity (kg)'].values[0]) + float(row['Kg_Terpakai'])
                                    df_stok_master.loc[batch_mask, 'Qty BAL'] = int(df_stok_master.loc[batch_mask, 'Qty BAL'].values[0]) + int(row['Bale_Terpakai'])
                                rev_row = {
                                    "Tipe_Pulp": str(row['Tipe_Pulp']), "Batch": str(row['Batch']), "Desc": str(row['Desc']), "Kode_Bahan": str(row['Kode_Bahan']),
                                    "Kg_Terpakai": -float(row['Kg_Terpakai']), "Bale_Terpakai": -int(row['Bale_Terpakai']),
                                    "No_PRO": str(rev_no_pro), "Customer": str(hist_pro['Customer'].iloc[0] if 'Customer' in hist_pro.columns else "Unknown"), "Kode_PEFC": str(hist_pro['Kode_PEFC'].iloc[0])
                                }
                                detail_reversals.append(rev_row)
                        save_stock_gspread(df_stok_master)
                        if detail_reversals:
                            append_log_gspread_strict(SHEET_DETAIL, detail_reversals)
                        customer_lama = hist_pro['Customer'].iloc[0] if 'Customer' in hist_pro.columns else "Unknown"
                        kode_pefc_lama = hist_pro['Kode_PEFC'].iloc[0]
                        log_reversal = [{
                            "No_PRO": str(rev_no_pro), "Tanggal": datetime.now().strftime('%Y-%m-%d'),
                            "Requester": "System (Dibatalkan)", "Customer": str(customer_lama),
                            "Kode_PEFC": str(kode_pefc_lama), "Tonase": -float(total_req_saat_ini)
                        }]
                        append_log_gspread_strict(SHEET_LOG, log_reversal)
                        st.success(f"✅ Transaksi **{rev_no_pro}** dibatalkan. Riwayat pembatalan (minus) berhasil dicatat ke sistem!")

        elif jenis_revisi.startswith("3."):
            aktual_tonase = st.number_input("Hasil Produksi Aktual Akhir (Kg)", min_value=1.0, value=float(total_req_saat_ini) if total_req_saat_ini > 0 else 1000.0, step=10.0)
            if st.button("🔄 Proses Penyesuaian", type="primary"):
                with st.spinner("⏳ Menyesuaikan aktual & menghitung ulang stok..."):
                    hist_pro = df_detail_all[df_detail_all['No_PRO'].astype(str) == str(rev_no_pro)]
                    if hist_pro.empty:
                        st.warning("No PRO tidak ditemukan.")
                    else:
                        selisih = aktual_tonase - total_req_saat_ini
                        if selisih == 0:
                            st.info("Angka aktual sama persis.")
                        elif selisih < 0:
                            refund_tonase = abs(selisih)
                            pulp_kembali = refund_tonase / 0.93
                            tot_nbkp = pd.to_numeric(hist_pro[hist_pro['Tipe_Pulp'] == 'NBKP']['Kg_Terpakai'], errors='coerce').sum()
                            tot_ebkp = pd.to_numeric(hist_pro[hist_pro['Tipe_Pulp'] == 'EBKP']['Kg_Terpakai'], errors='coerce').sum()
                            tot_pulp = tot_nbkp + tot_ebkp
                            rasio_nbkp = tot_nbkp / tot_pulp if tot_pulp > 0 else 0
                            rasio_ebkp = tot_ebkp / tot_pulp if tot_pulp > 0 else 0
                            
                            kembali_nbkp = pulp_kembali * rasio_nbkp
                            kembali_ebkp = pulp_kembali * rasio_ebkp
                            
                            for idx, row in hist_pro[hist_pro['Tipe_Pulp'] == 'NBKP'].iloc[::-1].iterrows():
                                if kembali_nbkp <= 0: break
                                amt = min(kembali_nbkp, float(row['Kg_Terpakai']))
                                bmask = df_stok_master['Batch'].astype(str) == str(row['Batch'])
                                if not df_stok_master[bmask].empty:
                                    conv = float(df_stok_master.loc[bmask, 'Conv (kg/bale)'].values[0])
                                    b_ref = math.floor(amt / conv) if conv > 0 else 0
                                    df_stok_master.loc[bmask, 'Quantity (kg)'] = float(df_stok_master.loc[bmask, 'Quantity (kg)'].values[0]) + amt
                                    df_stok_master.loc[bmask, 'Qty BAL'] = int(df_stok_master.loc[bmask, 'Qty BAL'].values[0]) + b_ref
                                kembali_nbkp -= amt
                                
                            for idx, row in hist_pro[hist_pro['Tipe_Pulp'] == 'EBKP'].iloc[::-1].iterrows():
                                if kembali_ebkp <= 0: break
                                amt = min(kembali_ebkp, float(row['Kg_Terpakai']))
                                bmask = df_stok_master['Batch'].astype(str) == str(row['Batch'])
                                if not df_stok_master[bmask].empty:
                                    conv = float(df_stok_master.loc[bmask, 'Conv (kg/bale)'].values[0])
                                    b_ref = math.floor(amt / conv) if conv > 0 else 0
                                    df_stok_master.loc[bmask, 'Quantity (kg)'] = float(df_stok_master.loc[bmask, 'Quantity (kg)'].values[0]) + amt
                                    df_stok_master.loc[bmask, 'Qty BAL'] = int(df_stok_master.loc[bmask, 'Qty BAL'].values[0]) + b_ref
                                kembali_ebkp -= amt
                                
                            save_stock_gspread(df_stok_master)
                            customer_lama = hist_pro['Customer'].iloc[0] if 'Customer' in hist_pro.columns else "Unknown"
                            kode_pefc_lama = hist_pro['Kode_PEFC'].iloc[0]
                            log_refund = [{
                                "No_PRO": str(rev_no_pro), "Tanggal": datetime.now().strftime('%Y-%m-%d'),
                                "Requester": "System (Refund)", "Customer": str(customer_lama),
                                "Kode_PEFC": str(kode_pefc_lama), "Tonase": -float(refund_tonase)
                            }]
                            append_log_gspread_strict(SHEET_LOG, log_refund)
                            st.success(f"✅ Refund aktual sebesar **{refund_tonase} Kg** berhasil dikembalikan.")

# ----------------- TAB 4: TERIMA STOK -----------------
with tab_stok_masuk:
    st.header("📦 Kedatangan Bahan Baku Baru (Online)")
    st.subheader("1. Input Manual (Satu Per Satu)")
    with st.form("form_tambah_stok_online"):
        col1, col2 = st.columns(2)
        with col1:
            in_tgl = st.date_input("Tanggal Kedatangan")
            in_tipe = st.selectbox("Material Type", ["NBKP", "EBKP", "BBKP", "LBKP"])
            in_fsc = st.text_input("FSC CODE (Kode Bahan)").upper()
            in_batch = st.text_input("Nomor Batch")
            in_desc = st.text_input("Batch Description")
        with col2:
            in_kg = st.number_input("Total Quantity (Kg)", min_value=0.0, step=100.0)
            in_bal = st.number_input("Total Bale", min_value=0, step=1)
            in_claim_bb = st.selectbox("Claim Bahan Baku", ["PEFC Certified", "Controlled Source", "FSC 100%", "FSC Mix Credit"])
            in_eudr = st.selectbox("Status EUDR", ["Bukan EUDR", "EUDR"])
        btn_submit_stok = st.form_submit_button("💾 Simpan Stok ke Google Sheets", type="primary")
        if btn_submit_stok and in_batch != "" and in_kg > 0:
            with st.spinner("⏳ Menyimpan data stok baru ke Google Sheets..."):
                try:
                    df_curr, _ = load_data_gspread()
                    new_row = {col: None for col in df_curr.columns}
                    new_row.update({
                        'Batch': in_batch, 'Batch Description': in_desc,
                        'Quantity (kg)': in_kg, 'Qty BAL': in_bal,
                        'Conv (kg/bale)': (in_kg / in_bal) if in_bal > 0 else 0,
                        'FSC CODE': in_fsc, 'Material type': in_tipe,
                        'Claim BB': in_claim_bb, 'Tgl Kedatangan': in_tgl.strftime('%Y-%m-%d'),
                        'CLAIM EUDR': "EUDR" if in_eudr == "EUDR" else None
                    })
                    df_updated = pd.concat([df_curr, pd.DataFrame([new_row])], ignore_index=True)
                    save_stock_gspread(df_updated)
                    st.success(f"🎉 Sukses! Stok **{in_batch}** berhasil ditambahkan.")
                except Exception as e:
                    st.error(f"❌ Gagal: {e}")

    st.markdown("---")
    st.subheader("2. Upload Massal (Banyak Data Sekaligus)")
    st.info("💡 Pastikan file Excel Anda memiliki urutan dan penamaan judul kolom yang sama persis dengan master stok di Google Sheets.")
    file_upload = st.file_uploader("Upload File Excel (.xlsx)", type=["xlsx"])
    if file_upload is not None:
        if st.button("🚀 Proses Upload Massal ke Google Sheets", type="primary"):
            with st.spinner("⏳ Membaca file Excel & mengirim data masal ke Google Sheets..."):
                try:
                    df_baru = pd.read_excel(file_upload)
                    df_curr, _ = load_data_gspread()
                    df_updated = pd.concat([df_curr, df_baru], ignore_index=True)
                    save_stock_gspread(df_updated)
                    st.success(f"🎉 Sukses! Sebanyak **{len(df_baru)} baris data stok** berhasil ditambahkan ke Google Sheets.")
                except Exception as e:
                    st.error(f"❌ Gagal memproses file Excel: {e}")

# ----------------- TAB 5: DASHBOARD STOK -----------------
with tab_dashboard:
    st.header("📊 Dashboard Sisa Stok Pulp Terkini (Google Sheets)")
    df_dashboard, _ = load_data_gspread()
    if not df_dashboard.empty:
        df_dashboard['Quantity (kg)'] = pd.to_numeric(df_dashboard['Quantity (kg)'], errors='coerce').fillna(0)
        df_dashboard['Qty BAL'] = pd.to_numeric(df_dashboard['Qty BAL'], errors='coerce').fillna(0)
        df_aktif = df_dashboard[df_dashboard['Quantity (kg)'] > 0].copy()
        def get_kategori(row):
            tipe = str(row['Material type']).strip() if pd.notna(row['Material type']) else 'UNKNOWN'
            claim = str(row['Claim BB']).strip() if pd.notna(row['Claim BB']) else 'No Claim'
            eudr = str(row['CLAIM EUDR']).strip().upper()
            label = f"{tipe} - {claim}"
            if eudr == 'EUDR': label += " (EUDR)"
            return label
        df_aktif['Kategori_Pulp'] = df_aktif.apply(get_kategori, axis=1)
        rekap_stok = df_aktif.groupby('Kategori_Pulp')[['Quantity (kg)', 'Qty BAL']].sum().reset_index()
        rekap_stok['Quantity (kg)'] = rekap_stok['Quantity (kg)'].apply(lambda x: f"{x:,.2f} Kg")
        rekap_stok['Qty BAL'] = rekap_stok['Qty BAL'].apply(lambda x: f"{int(x):,} Bal")
        st.dataframe(rekap_stok, use_container_width=True, hide_index=True)

# ----------------- TAB 6: ADMIN & RESET (DIPERBAIKI DENGAN AUTO-BACKUP) -----------------
with tab_admin:
    st.header("⚙️ Pengaturan & Reset (Online Google Sheets)")
    st.warning("Halaman ini mereset log transaksi online. Master stok Google Sheets TIDAK akan dihapus.")

    st.subheader("Langkah 1: Amankan Data (Backup)")
    st.info("Sangat disarankan untuk mendownload data riwayat saat ini sebelum Anda menghapusnya permanen.")
    
    df_log_utama_backup = get_all_log_gspread(SHEET_LOG)
    df_log_detail_backup = get_all_log_gspread(SHEET_DETAIL)
    
    if not df_log_utama_backup.empty or not df_log_detail_backup.empty:
        output_backup = io.BytesIO()
        with pd.ExcelWriter(output_backup, engine='openpyxl') as writer:
            if not df_log_utama_backup.empty: df_log_utama_backup.to_excel(writer, index=False, sheet_name='Summary_PRO')
            if not df_log_detail_backup.empty: df_log_detail_backup.to_excel(writer, index=False, sheet_name='Detail_Bahan_Baku')
        output_backup.seek(0)
        st.download_button(
            label="📥 Download File Backup Sekarang",
            data=output_backup,
            file_name=f"Backup_Sebelum_Reset_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.success("Log transaksi saat ini sudah kosong. Tidak ada data yang perlu dibackup.")

    st.markdown("---")
    
    st.subheader("Langkah 2: Eksekusi Reset")
    konfirmasi_reset = st.checkbox("Saya sudah mendownload backup & yakin ingin mereset log transaksi online.")
    
    if st.button("🚨 Eksekusi Reset Log Online", type="primary"):
        with st.spinner("⏳ Mengosongkan log transaksi & menata ulang judul kolom..."):
            if not konfirmasi_reset:
                st.error("Centang kotak konfirmasi terlebih dahulu!")
            else:
                try:
                    client = get_gsheet_client()
                    if client:
                        sheet_log = client.open(SHEET_LOG).sheet1
                        sheet_log.clear()
                        sheet_log.append_row(["No_PRO", "Tanggal", "Requester", "Customer", "Kode_PEFC", "Tonase"])
                        
                        sheet_detail = client.open(SHEET_DETAIL).sheet1
                        sheet_detail.clear()
                        sheet_detail.append_row(["Tipe_Pulp", "Batch", "Desc", "Kode_Bahan", "Kg_Terpakai", "Bale_Terpakai", "No_PRO", "Customer", "Kode_PEFC"])
                        
                        st.success("✅ Log transaksi berhasil dikosongkan secara aman (Judul kolom tetap dipertahankan).")
                except Exception as e:
                    st.error(f"❌ Gagal mereset: {e}")